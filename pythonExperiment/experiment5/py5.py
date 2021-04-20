import openpyxl
excel = openpyxl.load_workbook('resources/成绩.xlsx')
sheet1 = excel['Sheet1']
m = {}
lst = []
for row in range(2, sheet1.max_row + 1):
    key = sheet1.cell(row=row, column=1).value
    m[key] = sum([int(sheet1.cell(row=row, column=x).value)
                  for x in range(2, 5)])
    lst.append(key)


def cmp(key):
    return m[key]


lst.sort(key=cmp, reverse=True)

print('前三名为：')
for index in range(0, 3):
    print('{}:\t{}'.format(lst[index], m[lst[index]]))
